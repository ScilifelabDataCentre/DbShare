"Database HTML endpoints."

import copy
import csv
import hashlib
import io
import itertools
import json
import os
import os.path
import re
import shutil
import sqlite3
import stat
import tarfile
import tempfile
import urllib.parse

import flask
import openpyxl

import dbshare.system
import dbshare.table
import dbshare.query
import dbshare.user

from dbshare import constants
from dbshare import utils


TABLES_TABLE = {
    "name": constants.TABLES,
    "columns": [
        dict(name="name", type=constants.TEXT, primarykey=True),
        dict(name="schema", type=constants.TEXT, notnull=True),
    ],
}

INDEXES_TABLE = {
    "name": constants.INDEXES,
    "columns": [
        dict(name="name", type=constants.TEXT, primarykey=True),
        dict(name="schema", type=constants.TEXT, notnull=True),
    ],
}

VIEWS_TABLE = {
    "name": constants.VIEWS,
    "columns": [
        dict(name="name", type=constants.TEXT, primarykey=True),
        dict(name="schema", type=constants.TEXT, notnull=True),
    ],
}


blueprint = flask.Blueprint("db", __name__)


@blueprint.route("/<nameext:dbname>")
def display(dbname):
    "List the database tables, views and metadata."
    try:
        db = get_check_read(str(dbname), nrows=True)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))

    if dbname.ext in ("tar", "tar.gz", "tar.bz2"):
        try:
            mode = "w:" + dbname.ext.split(".")[1]
        except IndexError:
            mode = "w"
        outfile = io.BytesIO()
        tar = tarfile.open(fileobj=outfile, mode=mode)
        schemas = list(db["tables"].values()) + list(db["views"].values())
        for schema in schemas:
            columns = [c["name"] for c in schema["columns"]]
            sql = 'SELECT %s FROM "%s"' % (
                ",".join([f'"{c}"' for c in columns]),
                schema["name"],
            )
            writer = utils.CsvWriter(header=columns)
            try:
                cursor = utils.execute_timeout(get_cnx(db["name"]), sql)
            except SystemError:
                pass
            else:
                writer.write_rows(cursor)
                data = writer.getvalue().encode("utf-8")
                tarinfo = tarfile.TarInfo(name=f"{dbname}/{schema['name']}.csv")
                tarinfo.size = len(data)
                tar.addfile(tarinfo, io.BytesIO(data))
        tar.close()
        response = flask.make_response(outfile.getvalue())
        response.headers.set("Content-Type", constants.TAR_MIMETYPE)
        response.headers.set(
            "Content-Disposition", "attachment", filename=f"{dbname}.{dbname.ext}"
        )
        return response

    elif dbname.ext == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        tables = list(db["tables"].values())
        if tables:
            ws.title = tables[0]["name"]
            for table in tables[1:]:
                ws = wb.create_sheet(title=table["name"])
        for table in tables:
            ws = wb[table["name"]]
            columns = [c["name"] for c in table["columns"]]
            ws.append(columns)
            sql = 'SELECT %s FROM "%s"' % (
                ",".join([f'"{c}"' for c in columns]),
                table["name"],
            )
            try:
                cursor = utils.execute_timeout(get_cnx(db["name"]), sql)
            except SystemError:
                pass
            else:
                for row in cursor:
                    ws.append(list(row))
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            content = tmp.read()
        response = flask.make_response(content)
        response.headers.set("Content-Type", constants.XLSX_MIMETYPE)
        response.headers.set(
            "Content-Disposition", "attachment", filename=f"{dbname}.{dbname.ext}"
        )
        return response

    elif dbname.ext in (None, "html"):
        return flask.render_template(
            "db/display.html",
            db=db,
            title=db.get("title") or "Database {}".format(dbname),
            has_write_access=has_write_access(db),
            can_change_mode=has_write_access(db, check_mode=False),
        )

    else:
        flask.abort(http.client.NOT_ACCEPTABLE)


@blueprint.route("/", methods=["GET", "POST"])
@utils.login_required
def create():
    "Create a database."
    try:
        check_quota()
    except ValueError as error:
        utils.flash_error(error)
        return flask.redirect(
            flask.url_for("dbs.owner", username=flask.g.current_user["username"])
        )

    if utils.http_GET():
        return flask.render_template("db/create.html")

    elif utils.http_POST():
        try:
            with DbSaver() as saver:
                saver.set_name(flask.request.form["name"])
                saver.initialize()
                saver.set_title(flask.request.form.get("title"))
        except (KeyError, ValueError) as error:
            utils.flash_error(error)
            return flask.redirect(flask.url_for(".create"))
        return flask.redirect(flask.url_for(".display", dbname=saver.db["name"]))


@blueprint.route("/<name:dbname>/edit", methods=["GET", "POST", "DELETE"])
@utils.login_required
def edit(dbname):
    "Edit the database metadata. Or delete the database."
    try:
        db = get_check_write(dbname)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for(".display", dbname=dbname))

    if utils.http_GET():
        return flask.render_template("db/edit.html", db=db)

    elif utils.http_POST():
        try:
            with DbSaver(db) as saver:
                name = flask.request.form.get("name")
                if name:
                    saver.set_name(name)
                try:
                    saver.set_title(flask.request.form["title"])
                except KeyError:
                    pass
                try:
                    saver.set_description(flask.request.form["description"])
                except KeyError:
                    pass
        except (KeyError, ValueError) as error:
            utils.flash_error(error)
        return flask.redirect(flask.url_for(".display", dbname=db["name"]))

    elif utils.http_DELETE():
        delete_database(dbname)
        return flask.redirect(flask.url_for("dbs.owner", username=db["owner"]))


@blueprint.route("/<name:dbname>/logs")
def logs(dbname):
    "Display the logs for a database."
    try:
        db = get_check_read(dbname)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))
    sql = (
        "SELECT new, editor, remote_addr, user_agent, timestamp"
        " FROM dbs_logs WHERE name=? ORDER BY timestamp DESC"
    )
    logs = [
        {
            "new": json.loads(row[0]),
            "editor": row[1],
            "remote_addr": row[2],
            "user_agent": row[3],
            "timestamp": row[4],
        }
        for row in flask.g.syscnx.execute(sql, (db["name"],))
    ]
    return flask.render_template("db/logs.html", db=db, logs=logs)


@blueprint.route("/<name:dbname>/upload", methods=["GET", "POST"])
@utils.login_required
def upload(dbname):
    "Create a table from the data in a CSV file."
    try:
        check_quota()
        db = get_check_write(dbname)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for(".display", dbname=dbname))

    if utils.http_GET():
        return flask.render_template("db/upload.html", db=db)

    elif utils.http_POST():
        try:
            csvfile = flask.request.files["csvfile"]
            infile = io.StringIO(csvfile.read().decode("utf-8"))
            delimiter = flask.request.form.get("delimiter") or "comma"
            try:
                delimiter = flask.current_app.config["CSV_FILE_DELIMITERS"][delimiter][
                    "char"
                ]
            except KeyError:
                raise ValueError("invalid delimiter")
            try:
                tablename = flask.request.form["tablename"]
                if not tablename:
                    raise KeyError
            except KeyError:
                tablename = os.path.basename(csvfile.filename)
                tablename = os.path.splitext(tablename)[0]
            tablename = utils.name_cleaned(tablename)
            if utils.name_in_nocase(tablename, db["tables"]):
                raise ValueError("table name already in use")
            # Read the file; eliminate empty records.
            records = [r for r in csv.reader(infile, delimiter=delimiter) if r]
            if not records:
                raise ValueError("empty CSV file")
            # Change empty string items to None.
            for record in records:
                for i, item in enumerate(record):
                    if item == "":
                        record[i] = None
            has_header = utils.to_bool(flask.request.form.get("header"))
            with DbSaver(db) as saver:
                saver.create_table_load_records(
                    tablename, records, has_header=has_header
                )
            utils.flash_message(f"Loaded {len(records)} records.")
        except (ValueError, IndexError, sqlite3.Error) as error:
            utils.flash_error(error)
            return flask.redirect(
                flask.url_for(".upload", dbname=dbname, tablename=tablename)
            )
        return flask.redirect(
            flask.url_for("table.rows", dbname=dbname, tablename=tablename)
        )


@blueprint.route("/<name:dbname>/clone", methods=["GET", "POST"])
@utils.login_required
def clone(dbname):
    "Create a clone of the database."
    try:
        check_quota()
        db = get_check_read(dbname)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))

    if utils.http_GET():
        return flask.render_template("db/clone.html", db=db)

    elif utils.http_POST():
        try:
            with DbSaver() as saver:
                name = flask.request.form["name"]
                saver.set_name(name)
                saver.set_title(flask.request.form.get("title"))
                saver.set_description(flask.request.form.get("description"))
        except (KeyError, ValueError) as error:
            utils.flash_error(error)
            return flask.redirect(flask.url_for(".clone", dbname=dbname))
        shutil.copyfile(utils.get_dbpath(dbname), utils.get_dbpath(saver.db["name"]))
        db = get_db(name, complete=True)
        with DbSaver(db) as saver:
            saver.db["cloned"] = dbname  # Will show up in logs
        return flask.redirect(flask.url_for(".display", dbname=db["name"]))


@blueprint.route("/<name:dbname>/download")
def download(dbname):
    "Download the Sqlite3 database file."
    try:
        db = get_check_read(dbname)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))
    return flask.send_file(
        utils.get_dbpath(dbname),
        mimetype=constants.SQLITE3_MIMETYPE,
        as_attachment=True,
    )


@blueprint.route("/<name:dbname>/vacuum", methods=["POST"])
@utils.login_required
def vacuum(dbname):
    "Run VACUUM on the database. Also reset the table caches."
    utils.check_csrf_token()
    try:
        db = get_check_write(dbname)  # Do NOT allow if read-only (for hashes)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))
    try:
        # Reset the table caches.
        with DbSaver(db) as saver:
            for schema in db["tables"].values():
                saver.update_table(schema)
        get_cnx(db["name"], write=True).execute("VACUUM")
    except sqlite3.Error as error:
        utils.flash_error(error)
    return flask.redirect(flask.url_for(".display", dbname=db["name"]))


@blueprint.route("/<name:dbname>/analyze", methods=["POST"])
@utils.login_required
def analyze(dbname):
    "Run ANALYZE on the database."
    utils.check_csrf_token()
    try:
        db = get_check_write(dbname)  # Do NOT allow if read-only (for hashes)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))
    try:
        get_cnx(db["name"], write=True).execute("ANALYZE")
    except sqlite3.Error as error:
        utils.flash_error(error)
    return flask.redirect(flask.url_for(".display", dbname=db["name"]))


@blueprint.route("/<name:dbname>/public", methods=["POST"])
@utils.login_required
def public(dbname):
    "Set the database to public access."
    utils.check_csrf_token()
    try:
        db = get_check_write(dbname, check_mode=False)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))
    try:
        with DbSaver(db) as saver:
            saver.set_public(True)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
    else:
        utils.flash_message("Database set to public access.")
    return flask.redirect(flask.url_for(".display", dbname=db["name"]))


@blueprint.route("/<name:dbname>/private", methods=["POST"])
@utils.login_required
def private(dbname):
    "Set the database to private access."
    utils.check_csrf_token()
    try:
        db = get_check_write(dbname, check_mode=False)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))
    try:
        with DbSaver(db) as saver:
            saver.set_public(False)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
    else:
        utils.flash_message("Database access set to private.")
    return flask.redirect(flask.url_for(".display", dbname=db["name"]))


@blueprint.route("/<name:dbname>/readwrite", methods=["POST"])
@utils.login_required
def readwrite(dbname):
    "Set the database to read-write mode."
    utils.check_csrf_token()
    try:
        db = get_check_write(dbname, check_mode=False)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))
    if db["readonly"]:
        try:
            with DbSaver(db) as saver:
                saver.set_readonly(False)
        except (KeyError, ValueError) as error:
            utils.flash_error(error)
        else:
            utils.flash_message("Database set to read-write mode.")
    return flask.redirect(flask.url_for(".display", dbname=db["name"]))


@blueprint.route("/<name:dbname>/readonly", methods=["POST"])
@utils.login_required
def readonly(dbname):
    "Set the database to read-only mode. Compute content hashes."
    utils.check_csrf_token()
    try:
        db = get_check_write(dbname, check_mode=False)
    except (KeyError, ValueError) as error:
        utils.flash_error(error)
        return flask.redirect(flask.url_for("home"))
    if not db["readonly"]:
        try:
            with DbSaver(db) as saver:
                saver.set_readonly(True)
        except (KeyError, ValueError) as error:
            utils.flash_error(error)
        else:
            utils.flash_message("Database set to read-only mode.")
    return flask.redirect(flask.url_for(".display", dbname=db["name"]))


@blueprint.route("/<name:dbname>/owner", methods=["GET", "POST"])
@utils.admin_required
def owner(dbname):
    "Change owner of the database."
    db = get_db(dbname)
    if not db:
        utils.flash_error("No such database.")
        return flask.redirect(flask.url_for("home"))

    if utils.http_GET():
        return flask.render_template("db/owner.html", db=db)

    elif utils.http_POST():
        user = dbshare.user.get_user(flask.request.form["username"])
        if not user:
            utils.flash_error("No such user.")
            return flask.redirect(flask.url_for(".display", dbname=db["name"]))
        try:
            with DbSaver(db) as saver:
                saver.set_owner(user)
        except (KeyError, ValueError) as error:
            utils.flash_error(error)
        return flask.redirect(flask.url_for(".display", dbname=db["name"]))


class DbSaver:
    "Context handler to create, modify and save metadata for a database."

    def __init__(self, db=None):
        if db is None:
            self.db = {
                "owner": flask.g.current_user["username"],
                "public": False,
                "readonly": False,
                "hashes": {},
                "created": utils.get_time(),
            }
            self.old = {}
        else:
            self.db = db
            self.old = copy.deepcopy(db)

    @property
    def dbcnx(self):
        "Connection the Sqlite3 database itself."
        try:
            return self._dbcnx
        except AttributeError:
            # Don't close connection at exit; done externally to the context.
            self._dbcnx = get_cnx(self.db["name"], write=True)
            return self._dbcnx

    def __enter__(self):
        return self

    def __exit__(self, etyp, einst, etb):
        if etyp is not None:
            return False
        for key in ["name", "owner"]:
            if not self.db.get(key):
                raise ValueError(f"invalid db: {key} not set")
        self.db["modified"] = utils.get_time()
        cnx = utils.get_cnx(write=True)
        with cnx:
            # Update the existing database entry in system.
            if self.old:
                sql = (
                    "UPDATE dbs SET name=?, owner=?, title=?,"
                    "description=?, public=?, readonly=?, modified=?"
                    " WHERE name=?"
                )
                cnx.execute(
                    sql,
                    (
                        self.db["name"],
                        self.db["owner"],
                        self.db.get("title"),
                        self.db.get("description"),
                        bool(self.db["public"]),
                        bool(self.db["readonly"]),
                        self.db["modified"],
                        self.old["name"],
                    ),
                )
                # The Sqlite3 database file was renamed in 'set_name'.
                if self.old.get("name") != self.db["name"]:
                    # Fix entries in log records.
                    sql = "UPDATE dbs_logs SET name=? WHERE name=?"
                    cnx.execute(sql, (self.db["name"], self.old["name"]))
                    # No need to fix hash values: is (or at least, was)
                    # in read/write mode, so db has no hash values.
                # Insert hash values if newly computed.
                if not self.old["hashes"] and self.db["hashes"]:
                    sql = (
                        "INSERT INTO dbs_hashes (name, hashname, hashvalue)"
                        " VALUES (?, ?, ?)"
                    )
                    for hashname in self.db["hashes"]:
                        cnx.execute(
                            sql,
                            (self.db["name"], hashname, self.db["hashes"][hashname]),
                        )
                # Delete hash values if removed.
                elif self.old["hashes"] and not self.db["hashes"]:
                    sql = "DELETE FROM dbs_hashes WHERE name=?"
                    cnx.execute(sql, (self.db["name"],))

            # New database.
            else:
                # This actually creates the database file.
                self.dbcnx
                # Create the database entry in system.
                sql = (
                    "INSERT INTO dbs"
                    " (name, owner, title, description, public, readonly,"
                    "  created, modified) VALUES (?, ?, ?, ?, ?, ?, ?, ?)"
                )
                cnx.execute(
                    sql,
                    (
                        self.db["name"],
                        self.db["owner"],
                        self.db.get("title"),
                        self.db.get("description"),
                        bool(self.db["public"]),
                        bool(self.db["readonly"]),
                        self.db["created"],
                        self.db["modified"],
                    ),
                )
            # Add log entry
            new = {}
            for key, value in self.db.items():
                if value != self.old.get(key):
                    new[key] = value
            new.pop("modified")
            try:
                editor = flask.g.current_user["username"]
            except AttributeError:
                editor = None
            if flask.has_request_context():
                remote_addr = str(flask.request.remote_addr)
                user_agent = str(flask.request.user_agent)
            else:
                remote_addr = None
                user_agent = None
            sql = (
                "INSERT INTO dbs_logs (name, new, editor,"
                " remote_addr, user_agent, timestamp)"
                " VALUES (?, ?, ?, ?, ?, ?)"
            )
            cnx.execute(
                sql,
                (
                    self.db["name"],
                    json.dumps(new),
                    editor,
                    remote_addr,
                    user_agent,
                    utils.get_time(),
                ),
            )
        # Set the OS-level file permissions.
        if self.db["readonly"]:
            os.chmod(utils.get_dbpath(self.db["name"]), stat.S_IREAD)
        else:
            os.chmod(utils.get_dbpath(self.db["name"]), stat.S_IREAD | stat.S_IWRITE)

    def set_name(self, name, modify=False):
        """Set or change the database name.
        If 'modify' is True, then allow modifying the name to make it unique.
        Return the final name.
        Raise ValueError if name is invalid or already in use.
        """
        assert not hasattr(self, "_dbcnx")  # Must be done before any write ops.
        if name == self.db.get("name"):
            return
        if not constants.NAME_RX.match(name):
            raise ValueError("invalid database name")
        if modify:
            modified = name
            for n in range(1, 1000):  # Bail out if too many.
                if get_db(modified) is None:
                    name = modified
                    break
                modified = f"{name}-{n}"
        if get_db(name):
            raise ValueError("database name already in use")
        old_dbname = self.db.get("name")
        if old_dbname:
            # Rename the Sqlite3 file if the database already exists.
            os.rename(utils.get_dbpath(old_dbname), utils.get_dbpath(name))
            # The entries in the dbs_log will be fixed in '__exit__'
        self.db["name"] = name
        return self.db["name"]

    def set_title(self, title):
        "Set the database title."
        self.db["title"] = str(title) or None

    def set_owner(self, user):
        "Set the database owner."
        self.db["owner"] = user["username"]

    def set_description(self, description):
        "Set the database description."
        self.db["description"] = str(description) or None

    def set_public(self, access):
        "Set to public (True) or private (False) access."
        self.db["public"] = bool(access)

    def set_readonly(self, mode):
        """Set to 'readonly' (True) or 'readwrite' (False).
        If 'readonly', then compute the hash values, else remove them.
        """
        if self.db["readonly"] == mode:
            return
        self.db["readonly"] = self.readonly = mode
        if mode:
            hashes = {}
            for hashname in flask.current_app.config["CONTENT_HASHES"]:
                hashes[hashname] = hashlib.new(hashname)
            with open(utils.get_dbpath(self.db["name"]), "rb") as infile:
                data = infile.read(8192)
                while data:
                    for hash in hashes.values():
                        hash.update(data)
                    data = infile.read(8192)
            for hashname in hashes:
                hashes[hashname] = hashes[hashname].hexdigest()
            self.db["hashes"] = hashes
        else:
            self.db["hashes"] = {}

    def initialize(self):
        "Create the DbShare metadata tables and indexes if they do not exist."
        # Implicitly creates the file, or checks that it is an Sqlite3 file.
        sql = get_sql_create_table(TABLES_TABLE, if_not_exists=True)
        self.dbcnx.execute(sql)
        sql = get_sql_create_table(INDEXES_TABLE, if_not_exists=True)
        self.dbcnx.execute(sql)
        sql = get_sql_create_table(VIEWS_TABLE, if_not_exists=True)
        self.dbcnx.execute(sql)

    def create_table_load_records(self, tablename, records, has_header=True):
        """Create and load table from records (lists of data items).
        Infer table column types and constraints from records contents.
        Raises ValueError or sqlite3.Error if any problem.
        """
        # Column names from header, or make up.
        if has_header:
            header = records.pop(0)
            header = [utils.name_cleaned(n) for n in header]
            if len(header) != len(set(header)):
                raise ValueError("non-unique header column names")
        else:
            header = [f"column{i+1}" for i in range(len(records[0]))]

        # Infer column types and constraints.
        schema = {"name": tablename}
        schema["columns"] = [{"name": name} for name in header]
        try:
            for i, column in enumerate(schema["columns"]):
                type = None
                column["notnull"] = True

                # First attempt: integer
                for n, record in enumerate(records):
                    value = record[i]
                    if value is None:
                        column["notnull"] = False
                    elif isinstance(value, int):
                        pass
                    elif isinstance(value, str):
                        try:
                            int(value)
                        except (ValueError, TypeError):
                            break
                    else:
                        break
                else:
                    type = constants.INTEGER

                # Next attempt: float
                if type is None:
                    for n, record in enumerate(records):
                        value = record[i]
                        if value is None:
                            column["notnull"] = False
                        elif isinstance(value, (float, int)):
                            pass
                        elif isinstance(value, str):
                            try:
                                float(value)
                            except (ValueError, TypeError):
                                break
                        else:
                            break
                    else:
                        type = constants.REAL

                # Default: text
                if type is None:
                    column["type"] = constants.TEXT
                    if column["notnull"]:
                        for record in records:
                            value = record[i]
                            if value is None:
                                column["notnull"] = False
                                break
                else:
                    column["type"] = type
        except IndexError:
            raise ValueError(f"record {i+1} has too few items")

        # Create the table.
        self.add_table(schema)

        # Actually convert values in records.
        for i, column in enumerate(schema["columns"]):
            type = column["type"]
            if type == constants.INTEGER:
                for n, record in enumerate(records):
                    value = record[i]
                    if value is not None:
                        record[i] = int(value)
            elif type == constants.REAL:
                for n, record in enumerate(records):
                    value = record[i]
                    if value is not None:
                        record[i] = float(value)

        # Insert the data.
        sql = 'INSERT INTO "%s" (%s) VALUES (%s)' % (
            tablename,
            ",".join(['"%(name)s"' % c for c in schema["columns"]]),
            ",".join("?" * len(schema["columns"])),
        )
        with self.dbcnx:
            self.dbcnx.executemany(sql, records)
        self.update_table(schema)

    def add_table(self, schema, query=None, create=True):
        """Create the table in the database and add to the database definition.
        If 'query' is given, do 'CREATE TABLE AS', and fix up the schema.
        If 'create' is True, then actually create the table.
        Raises ValueError if there is a problem with the input schema data.
        Raises SystemError if the query is interrupted by time-out.
        """
        if not constants.NAME_RX.match(schema["name"]):
            raise ValueError("invalid table name")
        if utils.name_in_nocase(schema["name"], self.db["tables"]):
            raise ValueError("name is already in use for a table")
        if utils.name_in_nocase(schema["name"], self.db["views"]):
            raise ValueError("name is already in use for a view")
        if query:
            sql = dbshare.query.get_sql_statement(query)
            sql = 'CREATE TABLE "%s" AS %s' % (schema["name"], sql)
            utils.execute_timeout(self.dbcnx, sql)
            if not schema.get("description"):
                schema["description"] = sql
            sql = 'PRAGMA table_info("%s")' % schema["name"]
            schema["columns"] = []
            for row in self.dbcnx.execute(sql):
                column = {"name": row[1]}
                if row[2] == "INT":
                    column["type"] = constants.INTEGER
                else:
                    column["type"] = row[2]
                schema["columns"].append(column)
        elif create:
            sql = get_sql_create_table(schema)
            self.dbcnx.execute(sql)
        with self.dbcnx:
            sql = f"INSERT INTO {constants.TABLES} (name,schema) VALUES (?,?)"
            self.dbcnx.execute(sql, (schema["name"], json.dumps(schema)))
        self.update_table(schema)
        self.db["tables"][schema["name"]] = schema

    def add_table_column(self, schema, column):
        """Add the given column to the table described by the schema.
        Raise ValueError if there is any problem.
        """
        if not column["name"] or not constants.NAME_RX.match(column["name"]):
            raise ValueError("invalid column name")
        if utils.name_in_nocase(column["name"], [c["name"] for c in schema["columns"]]):
            raise ValueError("non-unique column name")
        if column["type"] not in constants.COLUMN_TYPES:
            raise ValueError("invalid column type")
        sql = (
            f'''ALTER TABLE "{schema['name']}"'''
            f""" ADD COLUMN "{column['name']}" {column['type']}"""
        )
        if column.get("notnull"):
            notnull = ["NOT NULL"]
            if column["type"] == constants.INTEGER:
                notnull.append("DEFAULT 0")
            elif column["type"] == constants.REAL:
                notnull.append("DEFAULT 0.0")
            elif column["type"] in (constants.TEXT, constants.BLOB):
                notnull.append("DEFAULT ''")
            sql += " " + " ".join(notnull)
        self.dbcnx.execute(sql)
        schema["columns"].append(column)
        self.update_table(schema)

    def update_table(self, schema, reset_cache=True):
        """Update the table with the new schema, resetting the cached items:
        1) Recompute the 'nrows' value.
        2) Remove any column statistics.
        """
        if reset_cache:
            sql = f'''SELECT COUNT(*) FROM "{schema['name']}"'''
            schema["nrows"] = self.dbcnx.execute(sql).fetchone()[0]
            for column in schema["columns"]:
                column.pop("statistics", None)
        with self.dbcnx:
            sql = f"UPDATE {constants.TABLES} SET schema=? WHERE name=?"
            self.dbcnx.execute(sql, (json.dumps(schema), schema["name"]))
        self.db["tables"][schema["name"]] = schema

    def empty_table(self, schema):
        "Empty the table; delete all rows."
        with self.dbcnx:
            sql = f'''DELETE FROM "{schema['name']}"'''
            self.dbcnx.execute(sql)
            self.update_table(schema)

    def delete_table(self, tablename):
        "Delete the table from the database and from the database definition."
        try:
            self.db["tables"].pop(tablename)
        except KeyError:
            raise ValueError("no such table in database")
        # Delete all indexes for this table.
        for indexname in list(self.db["indexes"]):
            self.delete_index(indexname)

        # Delete all views having this table as source.
        # Will recursively delete other dependent views.
        for view in list(self.db["views"].values()):
            if tablename in view["sources"]:
                # Need to catch KeyError, since recursion might
                # have deleted the view before we get here.
                try:
                    self.delete_view(view["name"])
                except KeyError:
                    pass
        with self.dbcnx:
            sql = 'DELETE FROM "%s" WHERE name=?' % constants.TABLES
            self.dbcnx.execute(sql, (tablename,))
        sql = 'DROP TABLE "%s"' % tablename
        self.dbcnx.execute(sql)
        sql = "VACUUM"
        self.dbcnx.execute(sql)

    def add_index(self, tablename, schema):
        "Create an index in the database and add to the database definition."
        if not utils.name_in_nocase(tablename, self.db["tables"]):
            raise ValueError(
                f"no such table {tablename}" f" for index {schema['name']}"
            )
        name = schema.get("name", "")
        if name:
            if utils.name_in_nocase(name, self.db["indexes"]):
                raise ValueError(f"index {name} already defined")
        else:
            prefix = f"_index_{tablename}_"
            ordinal = -1
            for ix in self.db["indexes"]:
                if ix.startswith(prefix):
                    try:
                        ordinal = max(ordinal, int(ix[len(prefix) :]))
                    except (ValueError, TypeError, IndexError):
                        pass
            name = prefix + str(ordinal + 1)
        schema["table"] = tablename
        schema["name"] = name
        sql = get_sql_create_index(tablename, schema)
        self.dbcnx.execute(sql)
        sql = "INSERT INTO %s (name, schema) VALUES (?, ?)" % constants.INDEXES
        with self.dbcnx:
            self.dbcnx.execute(sql, (schema["name"], json.dumps(schema)))
        self.db["indexes"][schema["name"]] = schema

    def delete_index(self, indexname):
        "Delete an index in the database."
        try:
            self.db["indexes"].pop(indexname)
        except KeyError:
            raise ValueError("no such index in database")
        with self.dbcnx:
            sql = "DELETE FROM %s WHERE name=?" % constants.INDEXES
            self.dbcnx.execute(sql, (indexname,))
        sql = 'DROP INDEX "%s"' % indexname
        self.dbcnx.execute(sql)

    def add_view(self, schema, create=True):
        """Create a view in the database and add to the database definition.
        If 'create' is True, then actually create the view.
        Raises ValueError if there is a problem with the input schema data.
        """
        if not constants.NAME_RX.match(schema["name"]):
            raise ValueError("invalid view name")
        if utils.name_in_nocase(schema["name"], self.db["tables"]):
            raise ValueError("name is already in use for a table")
        if utils.name_in_nocase(schema["name"], self.db["views"]):
            raise ValueError("name is already in use for a view")
        if create:
            sql = 'CREATE VIEW "%s" AS %s' % (
                schema["name"],
                dbshare.query.get_sql_statement(schema["query"]),
            )
            self.dbcnx.execute(sql)
        cursor = self.dbcnx.cursor()
        try:
            sql = 'PRAGMA table_info("%s")' % schema["name"]
            cursor.execute(sql)
        except sqlite3.Error:  # Invalid view
            sql = 'DROP VIEW "%s"' % schema["name"]
            cursor.execute(sql)
            raise ValueError("invalid view; maybe non-existent column?")
        # Source names considering quotes and disregarding AS part, if any.
        schema["sources"] = dbshare.query.get_from_sources(schema["query"]["from"])
        schema["columns"] = [{"name": row[1], "type": row[2]} for row in cursor]
        sql = "INSERT INTO %s (name, schema) VALUES (?,?)" % constants.VIEWS
        with self.dbcnx:
            self.dbcnx.execute(sql, (schema["name"], json.dumps(schema)))
        self.db["views"][schema["name"]] = schema

    def update_view(self, schema):
        "Update the view with the new schema."
        with self.dbcnx:
            sql = f"UPDATE {constants.VIEWS} SET schema=? WHERE name=?"
            self.dbcnx.execute(sql, (json.dumps(schema), schema["name"]))
        self.db["views"][schema["name"]] = schema

    def delete_view(self, viewname):
        "Delete a view in the database."
        try:
            self.db["views"].pop(viewname)
        except KeyError:
            raise ValueError("no such view in database")

        # Delete all views having this view as a source.
        # Will recursively delete other dependent views.
        for view in list(self.db["views"].values()):
            if viewname in view["sources"]:
                # Need to catch KeyError, since recursion might
                # have deleted the view before we get here.
                try:
                    self.delete_view(view["name"])
                except KeyError:
                    pass
        with self.dbcnx:
            sql = "DELETE FROM %s WHERE name=?" % constants.VIEWS
            self.dbcnx.execute(sql, (viewname,))
        sql = 'DROP VIEW "%s"' % viewname
        self.dbcnx.execute(sql)

    def check_metadata(self):
        """Check the validity of the metadata for the database.
        Return False if no metadata (i.e. not a DbShare file), else True.
        Raises ValueError or sqlite3.Error if any problem.
        """
        sql = f"SELECT COUNT(*) FROM {constants.TABLES}"
        if self.dbcnx.execute(sql).fetchone()[0] == 0:
            return False  # No metadata; skip.
        sql = f"SELECT name FROM {constants.TABLES}"
        tables1 = [r[0] for r in self.dbcnx.execute(sql)]
        sql = "SELECT name FROM sqlite_master WHERE type=?"
        tables2 = [r[0] for r in self.dbcnx.execute(sql, ("table",))]
        # Do not consider metadata tables and sqlite statistics tables, if any.
        tables2 = [n for n in tables2 if not n.startswith("_")]
        tables2 = [n for n in tables2 if not n.startswith("sqlite_")]
        if set(tables1) != set(tables2):
            raise ValueError("corrupt metadata in DbShare Sqlite3 file")
        # Does the index metatable exist?
        sql = f"SELECT name, schema FROM {constants.INDEXES}"
        self.dbcnx.execute(sql)
        # Does the views metatable exist?
        sql = f"SELECT name, schema FROM {constants.VIEWS}"
        self.dbcnx.execute(sql)
        return True

    def infer_metadata(self):
        "Infer and save the metadata for the database."
        cursor = self.dbcnx.cursor()
        # Get the table names.
        sql = "SELECT name FROM sqlite_master WHERE type=?"
        cursor.execute(sql, ("table",))
        tablenames = [
            row[0] for row in cursor if not row[0].startswith("_")
        ]  # Ignore metadata tables.
        # Check the DbShare validity of the table names.
        for tablename in tablenames:
            if not constants.NAME_RX.match(tablename):
                raise ValueError(f"invalid table name '{tablename}' for DbShare")
        # Infer the schema for the tables, and set the metadata.
        for tablename in tablenames:
            schema = {"name": tablename, "columns": []}
            sql = f'PRAGMA table_info("{tablename}")'
            cursor.execute(sql)
            for row in cursor:
                column = {
                    "name": row[1],
                    "type": row[2],
                    "notnull": bool(row[3]),
                    "primarykey": bool(row[5]),
                }
                schema["columns"].append(column)
            self.add_table(schema, create=False)
        # Get the views, attempt to parse their SQL definitions, and add.
        sql = "SELECT name, sql FROM sqlite_master WHERE type=?"
        cursor.execute(sql, ("view",))
        viewdata = [(row[0], row[1]) for row in cursor]
        for viewname, sql in viewdata:
            schema = {"name": viewname, "title": None, "description": None}
            schema["query"] = query = {}
            try:
                utils.lexer(sql)
                utils.lexer.get_expected("RESERVED", value="CREATE")
                utils.lexer.get_expected("RESERVED", value="VIEW")
                utils.lexer.get_expected("IDENTIFIER")
                utils.lexer.get_expected("RESERVED", value="AS")

                parts = utils.lexer.split_reserved(
                    ["SELECT", "FROM", "WHERE", "ORDER", "BY", "LIMIT", "OFFSET"]
                )

                query["select"] = "".join([t["raw"] for t in parts["SELECT"]]).strip()
                query["from"] = "".join([t["raw"] for t in parts["FROM"]]).strip()
                query["where"] = "".join([t["raw"] for t in parts["WHERE"]]).strip()
                try:
                    query["orderby"] = "".join([t["raw"] for t in parts["BY"]]).strip()
                except KeyError:
                    pass
                try:
                    query["limit"] = parts["LIMIT"][0]
                    query["offset"] = parts["OFFSET"][0]
                except KeyError:
                    query["limit"] = None
                    query["offset"] = None
                self.add_view(schema, create=False)
            except (KeyError, ValueError, IndexError, TypeError) as error:
                # Get rid of uninterpretable view.
                sql = f"DROP VIEW {viewname}"
                cursor.execute(sql)
        # Delete all indexes; currently not parsed and may interfere.
        sql = "SELECT name FROM sqlite_master WHERE type=?"
        cursor.execute(sql, ("index",))
        indexnames = [row[0] for row in cursor]
        # Do not attempt to delete Sqlite3 indexes.
        indexnames = [n for n in indexnames if not n.startswith("sqlite_autoindex")]
        indexnames = [n for n in indexnames if not n.startswith("_")]
        for name in indexnames:
            sql = f"DROP INDEX {name}"
            cursor.execute(sql)


def get_db(name, complete=False):
    """Return the database metadata for the given name.
    Return None if no such database.
    """
    cursor = flask.g.syscnx.cursor()
    sql = (
        "SELECT owner, title, description, public, readonly,"
        " created, modified FROM dbs WHERE name=?"
    )
    cursor.execute(sql, (name,))
    rows = cursor.fetchall()
    if len(rows) != 1:
        return None  # 'rowcount' does not work?!
    db = {"name": name}
    db.update(rows[0])
    db["public"] = bool(db["public"])
    db["readonly"] = bool(db["readonly"])
    db["size"] = os.path.getsize(utils.get_dbpath(name))
    db["hashes"] = {}
    sql = "SELECT hashname, hashvalue FROM dbs_hashes WHERE name=?"
    cursor.execute(sql, (name,))
    for row in cursor:
        db["hashes"][row[0]] = row[1]
    if complete:
        cursor = get_cnx(name).cursor()
        sql = "SELECT name, schema FROM %s" % constants.TABLES
        cursor.execute(sql)
        db["tables"] = dict([(row[0], json.loads(row[1])) for row in cursor])
        sql = "SELECT name, schema FROM %s" % constants.INDEXES
        cursor.execute(sql)
        db["indexes"] = dict([(row[0], json.loads(row[1])) for row in cursor])
        sql = "SELECT name, schema FROM %s" % constants.VIEWS
        cursor.execute(sql)
        db["views"] = dict([(row[0], json.loads(row[1])) for row in cursor])
    return db


def get_usage(username=None):
    "Return the number and total size of the databases for the user, or all."
    cursor = flask.g.syscnx.cursor()
    if username:
        sql = "SELECT name FROM dbs WHERE owner=?"
        cursor.execute(sql, (username,))
    else:
        sql = "SELECT name FROM dbs"
        cursor.execute(sql)
    rows = cursor.fetchall()
    return (len(rows), sum([os.path.getsize(utils.get_dbpath(row[0])) for row in rows]))


def check_quota(user=None, size=0):
    "Raise ValueError if the current user has exceeded her size quota."
    if user is None:
        user = flask.g.current_user
    quota = user["quota"]
    total_size = get_usage(user["username"])[1] + size
    if quota is not None and total_size > quota:
        raise ValueError("size quota exceeded; cannot add data")


def get_schema(db, sourcename):
    """Get the schema of the table or view.
    Add a member 'type' denoting which it is.
    Raise ValueError if no such table or view.
    """
    try:
        schema = db["tables"][sourcename]
        schema["type"] = constants.TABLE
    except KeyError:
        try:
            schema = db["views"][sourcename]
            schema["type"] = constants.VIEW
        except KeyError:
            raise ValueError("no such table/view")
    return schema


def get_sql_create_table(schema, if_not_exists=False):
    """Return SQL to create a table given by its schema.
    Raise ValueError if any problem.
    """
    if not schema.get("name"):
        raise ValueError("no table name defined")
    if not schema.get("columns"):
        raise ValueError("no columns defined")
    names = set()
    for column in schema["columns"]:
        if utils.name_in_nocase(column["name"], names):
            raise ValueError("column name %s repeated" % column["name"])
        if column["name"] == "rowid":
            raise ValueError("column name 'rowid' is reserved by the system")
        names.add(column["name"])
    # Collect columns forming primary key.
    primarykey = []
    for column in schema["columns"]:
        if column.get("primarykey"):
            primarykey.append(column["name"])
    # Column definitions, including column constraints.
    clauses = []
    for column in schema["columns"]:
        coldef = [f""""{column['name']}" {column['type']}"""]
        if column["name"] in primarykey:
            column["notnull"] = True
            if len(primarykey) == 1:
                coldef.append("PRIMARY KEY")
        if column.get("notnull"):
            coldef.append("NOT NULL")
        clauses.append(" ".join(coldef))
    # Primary key when more than one column.
    if len(primarykey) >= 2:
        clauses.append("PRIMARY KEY (%s)" % ",".join(['"%s"' for k in primarykey]))
    # Foreign keys.
    for foreignkey in schema.get("foreignkeys", []):
        clauses.append(
            'FOREIGN KEY (%s) REFERENCES "%s" (%s)'
            % (
                ",".join([f'"{c}"' for c in foreignkey["columns"]]),
                foreignkey["ref"],
                ",".join([f'"{c}"' for c in foreignkey["refcolumns"]]),
            )
        )
    sql = ["CREATE TABLE"]
    if if_not_exists:
        sql.append("IF NOT EXISTS")
    sql.append('"%s"' % schema["name"])
    sql.append("(%s)" % ", ".join(clauses))
    return " ".join(sql)


def get_sql_create_index(tablename, schema, if_not_exists=False):
    """Return SQL to create an index given by its schema.
    Raise ValueError if any problem.
    """
    if not schema.get("columns"):
        raise ValueError("no columns defined")
    if len(schema["columns"]) != len(set(schema["columns"])):
        raise ValueError("same column given more than once")
    sql = ["CREATE"]
    if schema.get("unique"):
        sql.append("UNIQUE")
    sql.append("INDEX")
    if if_not_exists:
        sql.append("IF NOT EXISTS")
    sql.append('"%s" ON "%s"' % (schema["name"], tablename))
    sql.append("(%s)" % ",".join([f'"{c}"' for c in schema["columns"]]))
    return " ".join(sql)


def get_cnx(dbname, write=False):
    """Get the connection for the database given by name.
    IMPORTANT: Only one connection to a non-system database can be open at any time!
    If the 'write' mode is wrong, then close and re-open.
    """
    try:
        if write == flask.g.dbwrite:
            return flask.g.dbcnx
        else:
            flask.g.dbcnx.close()
    except AttributeError:
        pass
    flask.g.dbcnx = utils.get_cnx(dbname, write=write)
    flask.g.dbname = dbname
    flask.g.dbwrite = write
    return flask.g.dbcnx


def has_read_access(db):
    "Does the current user (if any) have read access to the database?"
    if db["public"]:
        return True
    if not flask.g.current_user:
        return False
    if flask.g.is_admin:
        return True
    return flask.g.current_user["username"] == db["owner"]


def get_check_read(dbname, nrows=False, complete=True):
    """Get the database and check that the current user has read access.
    Optionally add nrows for each table and view.
    Raise KeyError if no such database.
    Raise ValueError if may not access.
    """
    db = get_db(dbname, complete=complete)
    if db is None:
        raise KeyError("no such database")
    if not has_read_access(db):
        raise ValueError("may not read the database")
    set_nrows(db, targets=nrows)
    return db


def has_write_access(db, check_mode=True):
    "Does the current user (if any) have write access to the database?"
    if not flask.g.current_user:
        return False
    if check_mode and db["readonly"]:
        return False
    if flask.g.is_admin:
        return True
    return flask.g.current_user["username"] == db["owner"]


def get_check_write(dbname, check_mode=True, nrows=False, complete=True):
    """Get the database and check that the current user has write access.
    Optionally add nrows for each table and view.
    Raise KeyError if no such database.
    Raise ValueError if may not access.
    """
    db = get_db(dbname, complete=complete)
    if db is None:
        raise KeyError("no such database")
    if not has_write_access(db, check_mode=check_mode):
        raise ValueError("may not write to the database")
    set_nrows(db, targets=nrows)
    return db


def set_nrows(db, targets):
    "Set the item 'nrows' for all or given tables and views of the database."
    if not targets:
        return
    if targets == True:
        targets = list(db["views"].values())
    else:
        targets = [get_schema(db, name) for name in targets]
    cnx = get_cnx(db["name"])
    for target in targets:
        try:
            utils.execute_timeout(cnx, _set_nrows, target=target)
        except SystemError:
            target["nrows"] = None


def _set_nrows(cnx, target):
    "Actually set the nrow values for the given target; executed with time-out."
    sql = 'SELECT COUNT(*) FROM "%s"' % target["name"]
    target["nrows"] = cnx.execute(sql).fetchone()[0]


def add_sqlite3_database(dbname, infile, size):
    """Add the Sqlite3 database file present in the given open file object.
    If the database has the metadata of a DbShare Sqlite3 database, check it.
    Else if the database appears to be a plain Sqlite3 database,
    infer the DbShare metadata from it by inspection.
    'size' is the size of the database file.
    Return the database dictionary.
    Raise ValueError if any problem.
    """
    try:
        check_quota(size=size)
        with DbSaver() as saver:
            dbname = saver.set_name(dbname, modify=True)
            with open(utils.get_dbpath(dbname), "wb") as outfile:
                outfile.write(infile.read())
            saver.initialize()
    except (ValueError, TypeError, OSError, IOError, sqlite3.Error) as error:
        raise ValueError(str(error))
    try:
        with DbSaver(get_db(dbname, complete=True)) as saver:  # Re-read db dict
            if not saver.check_metadata():
                saver.infer_metadata()
        return saver.db
    except (ValueError, TypeError, sqlite3.Error) as error:
        delete_database(dbname)
        raise ValueError(str(error))


def add_xlsx_database(dbname, infile, size):
    """Add the XLSX file workbook as a database.
    The worksheets are loaded as tables.
    'size' is the size of the XLSX file.
    Return the database dictionary.
    Raise ValueError if any problem.
    """
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx")
    tmp.write(infile.read())
    tmp.seek(0)
    try:
        wb = openpyxl.load_workbook(tmp.name)
        check_quota(size=size)
        with DbSaver() as saver:
            dbname = saver.set_name(dbname)
            saver.initialize()
        db = get_db(dbname, complete=True)
    except (ValueError, IOError) as error:
        raise ValueError(str(error))
    for sheet in wb:
        # Ensure the table name is unique.
        tname = utils.name_cleaned(sheet.title)
        tablename = tname
        count = 1
        while tablename in db["tables"]:
            count += 1
            tablename = f"{tname}{count}"
        records = list(sheet.values)
        # The header determines the number of columns;
        # clip off any trailing None values.
        for pos, item in enumerate(records[0]):
            if item is None:
                records[0] = records[0][:pos]
                break

        # Truncate records to same number of items as header; convert to lists.
        # Convert records from tuples to lists.
        length = len(records[0])
        records = [list(r[:length]) for r in records]
        with DbSaver(db) as saver:
            saver.create_table_load_records(tablename, records)
    return db


def delete_database(dbname):
    "Delete the database in the system database and from disk."
    cnx = utils.get_cnx(write=True)
    with cnx:
        sql = "DELETE FROM dbs_logs WHERE name=?"
        cnx.execute(sql, (dbname,))
        sql = "DELETE FROM dbs WHERE name=?"
        cnx.execute(sql, (dbname,))
    try:
        os.remove(utils.get_dbpath(dbname))
    except FileNotFoundError:
        pass
    
